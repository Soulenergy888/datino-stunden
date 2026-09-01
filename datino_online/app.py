#!/usr/bin/env python3
"""
DaTino Ristobar – Online Stundeneingabe
PostgreSQL (Railway) oder SQLite (lokal)
"""

import os, io, json
from datetime import date, datetime, timedelta
from flask import (Flask, request, redirect, url_for, render_template,
                   session, flash, send_file, jsonify)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'datino-geheim-2026')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'datino2026')
DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_PG = bool(DATABASE_URL)

MITARBEITER_INITIAL = [
    "Anxhela Kuci", "Parid Kuci", "Claudio Lorusso", "Alice Lumia",
    "Gayendra Kalhara Mcshane", "Hector Luis Nunez Pablo", "George Pavel",
    "Salvatore Petraroli", "Giuseppe Pilato", "Mihaela Plesa", "Ana-Maria Rus",
    "Ayad Ashraf Mahmoud Abdelghafar", "Anand Aditya Bhatt", "Salvatore Coluccio",
    "Jerome Gangabada Kanamge", "Georgiana Madalina Guta", "Obada Ioana-Veronica",
    "Achref Kadri", "Lennart Kramer",
]

# ── Datenbank ──────────────────────────────────────────────────────────────
def get_db():
    if USE_PG:
        import psycopg2
        import psycopg2.extras
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    else:
        import sqlite3
        conn = sqlite3.connect(os.environ.get('DB_PATH', 'stunden.db'))
        conn.row_factory = sqlite3.Row
        return conn

def db_execute(conn, sql, params=()):
    """Führt SQL aus (INSERT/UPDATE/DELETE/CREATE)."""
    if USE_PG:
        sql = sql.replace('?', '%s')
        sql = sql.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY')
        sql = sql.replace("datetime('now','localtime')", 'NOW()')
        sql = sql.replace('INSERT OR IGNORE', 'INSERT')
        # PostgreSQL: unique constraint verletzt → ignorieren
    cur = conn.cursor()
    try:
        cur.execute(sql, params)
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e

def db_fetchall(conn, sql, params=()):
    """Gibt Liste von dicts zurück."""
    if USE_PG:
        import psycopg2.extras
        sql = sql.replace('?', '%s')
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    else:
        cur = conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    return [dict(r) for r in rows]

def db_fetchone(conn, sql, params=()):
    rows = db_fetchall(conn, sql, params)
    return rows[0] if rows else None

def init_db():
    conn = get_db()
    try:
        if USE_PG:
            db_execute(conn, '''CREATE TABLE IF NOT EXISTS eintraege (
                id          SERIAL PRIMARY KEY,
                name        TEXT    NOT NULL,
                datum       TEXT    NOT NULL,
                beginn      TEXT    NOT NULL,
                ende        TEXT    NOT NULL,
                pause_min   INTEGER DEFAULT 0,
                bemerkungen TEXT    DEFAULT '',
                erstellt_am TIMESTAMP DEFAULT NOW()
            )''')
            db_execute(conn, '''CREATE TABLE IF NOT EXISTS mitarbeiter (
                id      SERIAL PRIMARY KEY,
                name    TEXT    NOT NULL UNIQUE,
                aktiv   INTEGER DEFAULT 1
            )''')
        else:
            db_execute(conn, '''CREATE TABLE IF NOT EXISTS eintraege (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT    NOT NULL,
                datum       TEXT    NOT NULL,
                beginn      TEXT    NOT NULL,
                ende        TEXT    NOT NULL,
                pause_min   INTEGER DEFAULT 0,
                bemerkungen TEXT    DEFAULT '',
                erstellt_am TEXT    DEFAULT (datetime('now','localtime'))
            )''')
            db_execute(conn, '''CREATE TABLE IF NOT EXISTS mitarbeiter (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                name    TEXT    NOT NULL UNIQUE,
                aktiv   INTEGER DEFAULT 1
            )''')

        count = db_fetchone(conn, 'SELECT COUNT(*) as c FROM mitarbeiter')['c']
        if count == 0:
            for name in MITARBEITER_INITIAL:
                try:
                    db_execute(conn, 'INSERT INTO mitarbeiter (name) VALUES (?)', (name,))
                except Exception:
                    pass
    finally:
        conn.close()

init_db()

def get_mitarbeiter():
    conn = get_db()
    try:
        rows = db_fetchall(conn,
            "SELECT name FROM mitarbeiter WHERE aktiv=1 "
            "ORDER BY LOWER(TRIM(SUBSTRING(name FROM POSITION(' ' IN name)+1)))"
            if USE_PG else
            "SELECT name FROM mitarbeiter WHERE aktiv=1 "
            "ORDER BY LOWER(TRIM(SUBSTR(name, INSTR(name,' ')+1)))"
        )
        return [r['name'] for r in rows]
    finally:
        conn.close()

def runde_viertelstunde(zeit):
    """Rundet 'HH:MM' auf die nächste Viertelstunde (00/15/30/45)."""
    try:
        h, m = map(int, zeit.split(':'))
        total = round((h*60 + m) / 15) * 15
        total %= 24*60
        return f'{total//60:02d}:{total%60:02d}'
    except Exception:
        return zeit

def netto_stunden(beginn, ende, pause_min):
    try:
        b = datetime.strptime(beginn, '%H:%M')
        e = datetime.strptime(ende,   '%H:%M')
        if e <= b: e += timedelta(hours=24)
        return round(max(0, (e-b).seconds/3600 - pause_min/60), 2)
    except Exception:
        return 0.0

# ── Mitarbeiter-Seite ──────────────────────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
def eingabe():
    success = error = None
    if request.method == 'POST':
        name      = request.form.get('name','').strip()
        datum     = request.form.get('datum','')
        beginn    = request.form.get('beginn','')
        ende      = request.form.get('ende','')
        pause_min = int(request.form.get('pause_min',0) or 0)
        bemerk    = request.form.get('bemerkungen','').strip()
        if not all([name,datum,beginn,ende]):
            error = 'Bitte alle Pflichtfelder ausfüllen.'
        else:
            beginn = runde_viertelstunde(beginn)
            ende   = runde_viertelstunde(ende)
            conn = get_db()
            try:
                db_execute(conn,
                    'INSERT INTO eintraege (name,datum,beginn,ende,pause_min,bemerkungen) VALUES (?,?,?,?,?,?)',
                    (name,datum,beginn,ende,pause_min,bemerk))
                netto = netto_stunden(beginn,ende,pause_min)
                success = f'✅ Gespeichert: {name}, {datum}, {beginn}–{ende} ({netto:.2f}h)'
            finally:
                conn.close()
    return render_template('eingabe.html', mitarbeiter=get_mitarbeiter(),
                           today=date.today().isoformat(), success=success, error=error)

# ── Admin Login ────────────────────────────────────────────────────────────
@app.route('/admin', methods=['GET','POST'])
def admin_login():
    if session.get('admin'): return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Falsches Passwort.')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('admin_login'))

# ── Admin Dashboard ────────────────────────────────────────────────────────
@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin'): return redirect(url_for('admin_login'))
    monat = request.args.get('monat', date.today().strftime('%Y-%m'))
    name  = request.args.get('name','')
    von   = monat+'-01'; bis = monat+'-31'
    conn  = get_db()
    try:
        if name:
            rows = db_fetchall(conn,
                'SELECT * FROM eintraege WHERE datum BETWEEN ? AND ? AND name=? ORDER BY datum,beginn',
                (von,bis,name))
        else:
            rows = db_fetchall(conn,
                'SELECT * FROM eintraege WHERE datum BETWEEN ? AND ? ORDER BY name,datum,beginn',
                (von,bis))
        monate = [r['substr'] if 'substr' in r else list(r.values())[0]
                  for r in db_fetchall(conn,
                    "SELECT DISTINCT SUBSTRING(datum,1,7) as substr FROM eintraege ORDER BY substr DESC"
                    if USE_PG else
                    "SELECT DISTINCT substr(datum,1,7) as substr FROM eintraege ORDER BY substr DESC")]
    finally:
        conn.close()
    eintraege = [{**r, 'netto': netto_stunden(r['beginn'],r['ende'],r['pause_min'])} for r in rows]
    return render_template('admin_dashboard.html',
                           eintraege=eintraege, gesamt=round(sum(e['netto'] for e in eintraege),2),
                           monat=monat, name_filter=name,
                           mitarbeiter=get_mitarbeiter(), monate=monate)

# ── Admin Edit ─────────────────────────────────────────────────────────────
@app.route('/admin/edit/<int:eid>', methods=['GET','POST'])
def admin_edit(eid):
    if not session.get('admin'): return redirect(url_for('admin_login'))
    conn = get_db()
    try:
        eintrag = db_fetchone(conn,'SELECT * FROM eintraege WHERE id=?',(eid,))
        if not eintrag:
            flash('Nicht gefunden.'); return redirect(url_for('admin_dashboard'))
        if request.method == 'POST':
            db_execute(conn,
                'UPDATE eintraege SET name=?,datum=?,beginn=?,ende=?,pause_min=?,bemerkungen=? WHERE id=?',
                (request.form.get('name'), request.form.get('datum'),
                 runde_viertelstunde(request.form.get('beginn','')),
                 runde_viertelstunde(request.form.get('ende','')),
                 int(request.form.get('pause_min',0) or 0),
                 request.form.get('bemerkungen',''), eid))
            flash('✅ Gespeichert.')
            return redirect(url_for('admin_dashboard'))
    finally:
        conn.close()
    return render_template('admin_edit.html', e=eintrag, mitarbeiter=get_mitarbeiter())

# ── Admin Delete ───────────────────────────────────────────────────────────
@app.route('/admin/delete/<int:eid>', methods=['POST'])
def admin_delete(eid):
    if not session.get('admin'): return redirect(url_for('admin_login'))
    conn = get_db()
    try:
        db_execute(conn,'DELETE FROM eintraege WHERE id=?',(eid,))
    finally:
        conn.close()
    flash('Gelöscht.')
    return redirect(request.referrer or url_for('admin_dashboard'))

# ── Admin Mitarbeiter ──────────────────────────────────────────────────────
@app.route('/admin/mitarbeiter', methods=['GET','POST'])
def admin_mitarbeiter():
    if not session.get('admin'): return redirect(url_for('admin_login'))
    if request.method == 'POST':
        aktion = request.form.get('aktion')
        conn = get_db()
        try:
            if aktion == 'add':
                name = request.form.get('name','').strip()
                if name:
                    try:
                        db_execute(conn,'INSERT INTO mitarbeiter (name) VALUES (?)',(name,))
                        flash(f'✅ {name} hinzugefügt.')
                    except Exception:
                        flash(f'⚠️ {name} existiert bereits.')
            elif aktion == 'toggle':
                db_execute(conn,'UPDATE mitarbeiter SET aktiv=1-aktiv WHERE id=?',(request.form.get('mid'),))
                flash('Aktualisiert.')
            elif aktion == 'delete':
                db_execute(conn,'DELETE FROM mitarbeiter WHERE id=?',(request.form.get('mid'),))
                flash('Gelöscht.')
        finally:
            conn.close()
        return redirect(url_for('admin_mitarbeiter'))

    conn = get_db()
    try:
        alle = db_fetchall(conn,
            "SELECT * FROM mitarbeiter ORDER BY LOWER(TRIM(SUBSTRING(name FROM POSITION(' ' IN name)+1)))"
            if USE_PG else
            "SELECT * FROM mitarbeiter ORDER BY LOWER(TRIM(SUBSTR(name, INSTR(name,' ')+1)))")
    finally:
        conn.close()
    return render_template('admin_mitarbeiter.html', mitarbeiter=alle)

# ── Admin Export ───────────────────────────────────────────────────────────
@app.route('/admin/export')
def admin_export():
    if not session.get('admin'): return redirect(url_for('admin_login'))
    monat = request.args.get('monat', date.today().strftime('%Y-%m'))
    name  = request.args.get('name','')
    von   = monat+'-01'; bis = monat+'-31'
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        conn = get_db()
        try:
            if name:
                rows = db_fetchall(conn,'SELECT * FROM eintraege WHERE datum BETWEEN ? AND ? AND name=? ORDER BY datum',(von,bis,name))
            else:
                rows = db_fetchall(conn,'SELECT * FROM eintraege WHERE datum BETWEEN ? AND ? ORDER BY name,datum',(von,bis))
        finally:
            conn.close()

        from openpyxl.worksheet.properties import PageSetupProperties
        from openpyxl.styles import Border, Side

        WD = ['Mo','Di','Mi','Do','Fr','Sa','So']
        MONATE = ['','Januar','Februar','März','April','Mai','Juni','Juli',
                  'August','September','Oktober','November','Dezember']
        jahr, mon = int(monat[:4]), int(monat[5:7])
        monat_name = f'{MONATE[mon]} {jahr}'

        def fmt_aufz(v):
            if isinstance(v, datetime): return v.strftime('%d.%m.%Y')
            s = str(v or '')[:10]
            try:    return datetime.strptime(s,'%Y-%m-%d').strftime('%d.%m.%Y')
            except Exception: return s

        # Einträge nach Mitarbeiter gruppieren
        von_mitarbeiter = {}
        for r in rows:
            von_mitarbeiter.setdefault(r['name'], []).append(r)
        if name:
            namen = [name]
        else:
            namen = sorted(von_mitarbeiter.keys(),
                           key=lambda n: n.split(' ')[-1].lower()) or ['Alle']

        thin = Side(style='thin', color='999999')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        kopf_fill = PatternFill('solid', fgColor='1A3A5C')

        def std_min(dezimal):
            """9.58 -> '9:35'"""
            m = int(round(dezimal*60))
            return f'{m//60}:{m%60:02d}'

        # Format wie Stundenliste_Monatsvorlage: Blatt 'Stundenliste',
        # B5 = Datum, Tag 1 -> Zeile 8 ... Tag 31 -> Zeile 38, Summe Zeile 39,
        # Zeiten als Excel-Zeitwerte, Bemerkungen in Spalte G.
        # So kann 'DaTino Stundenliste zu EA' die Datei direkt einlesen.
        import calendar
        from datetime import time as dt_time
        headers = ['Kalendertag','Beginn','Pause','Ende','Dauer (Std:Min)',
                   'aufgezeichnet am','Bemerkungen','Dezimal']
        tage_im_monat = calendar.monthrange(jahr, mon)[1]

        def mins(t):
            h,m = map(int, str(t).split(':')); return h*60+m

        def baue_stundenliste(mitarbeiter_name, eintr):
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = 'Stundenliste'
            # Seite: A4 Hochformat, alles auf 1 Seite
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.paperSize   = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth  = 1
            ws.page_setup.fitToHeight = 1
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.page_margins.left = ws.page_margins.right = 0.4
            ws.page_margins.top  = ws.page_margins.bottom = 0.5
            ws.print_options.horizontalCentered = True

            # Kopfbereich
            ws.merge_cells('A1:H1')
            ws['A1'] = 'Dokumentation der täglichen Arbeitszeit'
            ws['A1'].font = Font(bold=True, size=13)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A3'] = 'Firma:'; ws['A3'].font = Font(bold=True)
            ws.merge_cells('B3:H3'); ws['B3'] = 'Ristobar DaTino · Gesandtenstraße 18, 93047 Regensburg'
            ws['A4'] = 'Name des Mitarbeiters:'; ws['A4'].font = Font(bold=True)
            ws.merge_cells('B4:D4'); ws['B4'] = mitarbeiter_name
            ws['E4'] = 'Pers.-Nr.:'; ws['E4'].font = Font(bold=True)
            ws.merge_cells('F4:H4')
            ws['A5'] = 'Monat/Jahr:'; ws['A5'].font = Font(bold=True)
            ws.merge_cells('B5:H5')
            ws['B5'] = datetime(jahr, mon, 1)
            ws['B5'].number_format = 'MMMM YYYY'

            # Tabellenkopf (Zeile 7)
            for col,h in enumerate(headers,1):
                c = ws.cell(row=7,column=col,value=h)
                c.font = Font(bold=True,color='FFFFFF')
                c.fill = kopf_fill
                c.alignment = Alignment(horizontal='center', wrap_text=True)
                c.border = border

            # Einträge pro Kalendertag gruppieren (Splitschichten -> eine Zeile)
            pro_tag = {}
            for r in eintr:
                pro_tag.setdefault(str(r['datum'])[:10], []).append(r)

            for day in range(1, 32):
                zeile = day + 7                     # Tag 1 -> Zeile 8
                for col in range(1,9):
                    c = ws.cell(row=zeile,column=col)
                    c.border = border
                    if col <= 6: c.alignment = Alignment(horizontal='center')
                if day > tage_im_monat: continue
                d = date(jahr, mon, day)
                ws.cell(row=zeile,column=1,value=f'{WD[d.weekday()]}. {day:02d}')
                # Formeln: Dauer & Dezimal rechnen automatisch nach, wenn
                # Beginn/Pause/Ende in Excel geändert oder ergänzt werden
                z = zeile
                cE = ws.cell(row=z,column=5,value=(
                    f'=IF(OR(B{z}="",D{z}=""),"",'
                    f'IF(D{z}<B{z},D{z}+1-B{z},D{z}-B{z})-C{z})'))
                cE.number_format = 'h:mm'
                cH = ws.cell(row=z,column=8,value=f'=IF(E{z}="","",ROUND(E{z}*24,2))')
                cH.number_format = '0.00'
                liste = sorted(pro_tag.get(f'{jahr}-{mon:02d}-{day:02d}', []),
                               key=lambda r: r['beginn'])
                if not liste: continue
                beginn = liste[0]['beginn']; ende = liste[-1]['ende']
                pause  = sum(int(r['pause_min'] or 0) for r in liste)
                for prev,nxt in zip(liste, liste[1:]):
                    luecke = mins(nxt['beginn']) - mins(prev['ende'])
                    if luecke > 0: pause += luecke
                bem = ' / '.join(dict.fromkeys(
                    (r.get('bemerkungen') or '').strip()
                    for r in liste if (r.get('bemerkungen') or '').strip()))
                bh,bm = map(int, beginn.split(':')); eh,em = map(int, ende.split(':'))
                cb = ws.cell(row=zeile,column=2,value=dt_time(bh,bm)); cb.number_format='HH:MM'
                cp = ws.cell(row=zeile,column=3,value=dt_time(pause//60,pause%60)); cp.number_format='H:MM'
                ce = ws.cell(row=zeile,column=4,value=dt_time(eh,em)); ce.number_format='HH:MM'
                ws.cell(row=zeile,column=6,value=fmt_aufz(liste[0].get('erstellt_am')))
                ws.cell(row=zeile,column=7,value=bem)

            # Summenzeile (Zeile 39) – Formeln, rechnen bei Änderungen mit
            ws.merge_cells('A39:D39')
            cs = ws.cell(row=39,column=1,value='Summe:')
            cs.font = Font(bold=True); cs.alignment = Alignment(horizontal='right')
            ch = ws.cell(row=39,column=5,value='=SUM(E8:E38)')
            ch.number_format = '[h]:mm'
            ch.font = Font(bold=True); ch.alignment = Alignment(horizontal='center')
            cg = ws.cell(row=39,column=8,value='=ROUND(SUM(H8:H38),2)')
            cg.number_format = '0.00'
            cg.font = Font(bold=True); cg.alignment = Alignment(horizontal='center')
            for col in range(1,9):
                ws.cell(row=39,column=col).border = border

            for col,w in zip('ABCDEFGH',[11,9,8,9,12,14,20,9]):
                ws.column_dimensions[col].width = w
            return wb

        if name:
            # Ein Mitarbeiter -> eine Datei (Blatt 'Stundenliste')
            wb = baue_stundenliste(name, von_mitarbeiter.get(name, []))
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            fname = f'{name.replace(" ","_")}_{monat}.xlsx'
            return send_file(buf,as_attachment=True,download_name=fname,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            # Alle Mitarbeiter -> ZIP mit einer Datei pro Mitarbeiter
            if not von_mitarbeiter:
                flash('Keine Einträge für den ausgewählten Monat.')
                return redirect(url_for('admin_dashboard'))
            import zipfile
            zbuf = io.BytesIO()
            with zipfile.ZipFile(zbuf,'w',zipfile.ZIP_DEFLATED) as zf:
                for mn in namen:
                    wb = baue_stundenliste(mn, von_mitarbeiter[mn])
                    b = io.BytesIO(); wb.save(b)
                    zf.writestr(f'{mn.replace(" ","_")}.xlsx', b.getvalue())
            zbuf.seek(0)
            return send_file(zbuf,as_attachment=True,
                            download_name=f'DaTino_Stundenlisten_{monat}.zip',
                            mimetype='application/zip')
    except Exception as e:
        flash(f'Fehler: {e}'); return redirect(url_for('admin_dashboard'))

@app.route('/admin/json')
def admin_json():
    if not session.get('admin'): return jsonify({'error':'Nicht angemeldet'}),401
    monat = request.args.get('monat',date.today().strftime('%Y-%m'))
    name  = request.args.get('name','')
    von   = monat+'-01'; bis = monat+'-31'
    conn  = get_db()
    try:
        q = 'SELECT * FROM eintraege WHERE datum BETWEEN ? AND ?'
        p = [von,bis]
        if name: q+=' AND name=?'; p.append(name)
        rows = db_fetchall(conn,q+' ORDER BY name,datum',p)
    finally:
        conn.close()
    for r in rows:
        r['netto'] = netto_stunden(r['beginn'],r['ende'],r['pause_min'])
        if isinstance(r.get('erstellt_am'), datetime):
            r['erstellt_am'] = r['erstellt_am'].isoformat()
    return jsonify(rows)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
